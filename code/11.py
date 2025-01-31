import openpyxl
from openpyxl import Workbook

# 打开原始Excel文件
wb = openpyxl.load_workbook('D:\\桌面\\2025.1\\英文\\try.xlsx')
ws = wb.active  # 获取第一个sheet

# 创建一个新的工作簿
new_wb = Workbook()

# 获取年份列表
years = list(range(2009, 2014))

# 遍历每个年份，创建新的sheet并填充数据
for year in years:
    new_ws = new_wb.create_sheet(title=f'{year}')  # 创建新sheet
    # 获取该年份对应的数据列索引
    cols = [i for i, cell in enumerate(ws[2], start=1) if cell.value == year]
    # 复制标题行和数据行，同时删除空列
    for col_idx, col in enumerate(cols, start=1):
        # 复制标题行
        new_ws.cell(row=1, column=col_idx).value = ws.cell(row=1, column=col).value
        # 复制数据行
        for row in range(3, ws.max_row + 1):
            new_ws.cell(row=row-1, column=col_idx).value = ws.cell(row=row, column=col).value

# 删除默认创建的sheet
new_wb.remove(new_wb.active)

# 保存新的Excel文件
new_wb.save('D:\\桌面\\2025.1\\英文\\split_data.xlsx')